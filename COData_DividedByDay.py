#Defining (back to (Excel_CO2019) file for more details
import openpyxl
import os
from openpyxl.utils import get_column_letter
os.chdir("C:\\Users\\mmnnm\\Desktop\\PY_Master")

#Open a file in py and import CO data without the NULL DATA 

#The source Data
workbook = openpyxl.load_workbook('MoveTheDate_TheRealData_DeleteTheNULL.xlsx')
sheet = workbook['CO2019']
#NE: new excel
NE = openpyxl.Workbook()
#Activate the sheet in the new excel file
NEws = NE.active
#Name the sheet
NEws.title = 'CO2019'
#Headings
NEws.append([ sheet ['A6'].value , sheet ['B4'].value , sheet ['C4'].value , sheet ['D4'].value , sheet ['E4'].value , sheet ['F4'].value , sheet ['G4'].value , sheet ['H4'].value , sheet ['I4'].value , sheet ['J4'].value ])


#["%.2d" % day for day in range(1,24)]

##x = NEws['A7'].value
##Day1 = NE.create_sheet ('Day1')
##Day2 = NE.create_sheet ('Day2')
##for row in range (7,50):
##    char = get_column_letter (1)
##    cell = char + str (row)
##    x = str(sheet[cell].value)
##    if x.startswith("'2019-01-01") or x.startswith ("'01/01/2019"):
##        Day1 [cell] = x
##    else:
##        Day2[cell] = x


ListDay = ["%.2d" % i for i in range(32)] #List ['01','02',...]
ListMonth = ["%.2d" % i for i in range(13)]

for D in range (1,32):
    x = NE.create_sheet(str(D))


for row in range (9,50):
    for m in range(1,13):
        for d in range (1,32):
            char = get_column_letter (1)
            cell = char + str (row)
            x = str(sheet[cell].value)
            month = ListMonth[m]
            day = ListDay[d]
            if x.startswith("'2019-"+month+"-"+day):
                sheetnames = NE.sheetnames
                NE.active = int(sheetnames[d])
                NE.active[cell] = x
            else:
                continue
        
        

NE.save ('DividedByDay.xlsx')




