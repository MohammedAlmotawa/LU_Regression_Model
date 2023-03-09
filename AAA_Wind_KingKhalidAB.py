import openpyxl
from openpyxl.utils import get_column_letter

#Source Data
workbook = openpyxl.load_workbook('KingKhalidAB.xlsx')
sheet = workbook['Sheet']

#NE: new excel
NE = openpyxl.Workbook()


# Creating list for months. Ex: List ['00','01','02',...]
ListMonth = ["%.2d" % i for i in range(1,13)]

#Creating sheets for each month
for M in range (1,13):
    NE.create_sheet(str(M))
    

#Add the same Headings as the original file for each sheeat
for m in range(1,13):
    SW1 = NE.active
    SW1 = NE[str(m)]
    SW1.append([ sheet ['A1'].value , sheet ['B1'].value , sheet ['C1'].value , sheet ['D1'].value , sheet ['E1'].value , sheet ['F1'].value , sheet ['G1'].value , sheet ['H1'].value , sheet ['I1'].value , sheet ['J1'].value, sheet ['K1'].value , sheet ['L1'].value , sheet ['M1'].value , sheet ['N1'].value , sheet ['O1'].value , sheet ['P1'].value , sheet ['Q1'].value , sheet ['R1'].value , sheet ['S1'].value , sheet ['T1'].value, sheet ['U1'].value , sheet ['V1'].value , sheet ['W1'].value , sheet ['X1'].value , sheet ['Y1'].value , sheet ['Z1'].value , sheet ['AA1'].value , sheet ['AB1'].value , sheet ['AC1'].value , sheet ['AD1'].value, sheet ['AE1'].value , sheet ['AF1'].value , sheet ['AG1'].value , sheet ['AH1'].value , sheet ['AI1'].value ])



#Move every measurment based on the measurment month
for row in range (2,8900):
    for m in range(12):
        cell = 'E' + str (row)
        x = str(sheet[cell].value)
        month = ListMonth[m]
        if x.startswith("2019-"+month):
            SW2 = NE.active
            SW2 = NE[str(m+1)]
            SW2.append([ sheet ['A'+str(row)].value , sheet ['B'+str(row)].value , sheet ['C'+str(row)].value , sheet ['D'+str(row)].value , sheet ['E'+str(row)].value , sheet ['F'+str(row)].value , sheet ['G'+str(row)].value , sheet ['H'+str(row)].value , sheet ['I'+str(row)].value , sheet ['J'+str(row)].value, sheet ['K'+str(row)].value , sheet ['L'+str(row)].value , sheet ['M'+str(row)].value , sheet ['N'+str(row)].value , sheet ['O'+str(row)].value , sheet ['P'+str(row)].value , sheet ['Q'+str(row)].value , sheet ['R'+str(row)].value , sheet ['S'+str(row)].value , sheet ['T'+str(row)].value, sheet ['U'+str(row)].value , sheet ['V'+str(row)].value , sheet ['W'+str(row)].value , sheet ['X'+str(row)].value , sheet ['Y'+str(row)].value , sheet ['Z'+str(row)].value , sheet ['AA'+str(row)].value , sheet ['AB'+str(row)].value , sheet ['AC'+str(row)].value , sheet ['AD'+str(row)].value, sheet ['AE'+str(row)].value , sheet ['AF'+str(row)].value , sheet ['AG'+str(row)].value , sheet ['AH'+str(row)].value , sheet ['AI'+str(row)].value ])

        else:
            continue
        
#Delete the first sheet
del NE['Sheet']

#Save the file        
NE.save ('KingKhalidAB_ByMonth.xlsx')

