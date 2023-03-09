
from re import X
import openpyxl

#The source Data
WORKBOOK1 = openpyxl.load_workbook('AACO_DividedByMonth.xlsx')
SHEET1 = WORKBOOK1['1']
WORKBOOK2 = openpyxl.load_workbook('Abha_ByMonth.xlsx')
SHEET2 = WORKBOOK2['1']
WORKBOOK3 = openpyxl.load_workbook('Alahsa_ByMonth.xlsx')
SHEET3 = WORKBOOK3['1']
WORKBOOK4 = openpyxl.load_workbook('Albaha_ByMonth.xlsx')
SHEET4 = WORKBOOK4['1']
WORKBOOK5 = openpyxl.load_workbook('Albaha_ByMonth.xlsx')
SHEET5 = WORKBOOK5['1']
WORKBOOK6 = openpyxl.load_workbook('Aljouf_ByMonth.xlsx')
SHEET6 = WORKBOOK6['1']
WORKBOOK7 = openpyxl.load_workbook('Arar_ByMonth.xlsx')
SHEET7 = WORKBOOK7['1']

#NE: new excel
NE = openpyxl.Workbook()

# Creating lists for days and months. Ex: List ['01','02',...]
ListDay = ["%.2d" % i for i in range(1,32)] 
ListMonth = ["%.2d" % i for i in range(1,13)]

#Creating sheets for each day
for D in range (1,32):
    NE.create_sheet(str(D))

#Add the same Headings as the original file for each sheeat
#for d in range(1,32):
#    SW1 = NE.active
#    SW1 = NE[str(d)]
#    SW1.append([ SHEET1 ['A1'].value , SHEET1 ['B1'].value , SHEET1 ['C1'].value , SHEET1 ['D1'].value , SHEET1 ['E1'].value , SHEET1 ['F1'].value , SHEET1 ['G1'].value , SHEET1 ['H1'].value , SHEET1 ['I1'].value , SHEET1 ['J1'].value ])
#    SW1.append([ SHEET2 ['A1'].value , SHEET2 ['B1'].value , SHEET2 ['C1'].value , SHEET2 ['D1'].value , SHEET2 ['E1'].value , SHEET2 ['F1'].value , SHEET2 ['G1'].value , SHEET2 ['H1'].value , SHEET2 ['I1'].value , SHEET2 ['J1'].value, SHEET2 ['K1'].value , SHEET2 ['L1'].value , SHEET2 ['M1'].value , SHEET2 ['N1'].value , SHEET2 ['O1'].value , SHEET2 ['P1'].value , SHEET2 ['Q1'].value , SHEET2 ['R1'].value , SHEET2 ['S1'].value , SHEET2 ['T1'].value, SHEET2 ['U1'].value , SHEET2 ['V1'].value , SHEET2 ['W1'].value , SHEET2 ['X1'].value , SHEET2 ['Y1'].value , SHEET2 ['Z1'].value , SHEET2 ['AA1'].value , SHEET2 ['AB1'].value , SHEET2 ['AC1'].value , SHEET2 ['AD1'].value, SHEET2 ['AE1'].value , SHEET2 ['AF1'].value , SHEET2 ['AG1'].value , SHEET2 ['AH1'].value , SHEET2 ['AI1'].value ])


#Move every measurment based on the measurment month
for row in range (2,30): #CO #527
    cell = 'A' + str (row)
    X = str(SHEET1[cell].value)
    for d in range (31): #Day
        day = ListDay[d]
        for wind in range (2,30): #744
            Y = str(SHEET2['E'+str(wind)].value)
            Y1 = Y.startswith("2019-01-"+day+" ") #Wind
            X1 = X.startswith("'2019-01-"+day+" ") #CO
            if  X1 == Y1:
                SW2 = NE.active
                SW2 = NE[str(d+1)]
                #SW2.append([ SHEET2['A'+str(row)].value , SHEET2['B'+str(row)].value , SHEET2['C'+str(row)].value , SHEET2['D'+str(row)].value , SHEET2['E'+str(row)].value , SHEET2['F'+str(row)].value , SHEET2['G'+str(row)].value , SHEET2['H'+str(row)].value , SHEET2['I'+str(row)].value , SHEET2['J'+str(row)].value, SHEET2['K'+str(row)].value , SHEET2['L'+str(row)].value , SHEET2['M'+str(row)].value , SHEET2['N'+str(row)].value , SHEET2['O'+str(row)].value , SHEET2['P'+str(row)].value , SHEET2['Q'+str(row)].value , SHEET2['R'+str(row)].value , SHEET2['S'+str(row)].value , SHEET2['T'+str(row)].value, SHEET2['U'+str(row)].value , SHEET2['V'+str(row)].value , SHEET2['W'+str(row)].value , SHEET2['X'+str(row)].value , SHEET2['Y'+str(row)].value , SHEET2['Z'+str(row)].value , SHEET2['AA'+str(row)].value , SHEET2['AB'+str(row)].value , SHEET2['AC'+str(row)].value , SHEET2['AD'+str(row)].value, SHEET2['AE'+str(row)].value , SHEET2['AF'+str(row)].value , SHEET2['AG'+str(row)].value , SHEET2['AH'+str(row)].value , SHEET2['AI'+str(row)].value ])
                #SW2.append([SHEET1['A'+str(row)].value, SHEET1['B'+str(row)].value, SHEET1['C'+str(row)].value, SHEET1['D'+str(row)].value, SHEET1['E'+str(row)].value, SHEET1['F'+str(row)].value, SHEET1['G'+str(row)].value, SHEET1['H'+str(row)].value, SHEET1['I'+str(row)].value, SHEET1['J'+str(row)].value])
            
            else:
                continue


#Delete the first sheet
del NE['Sheet']

#Save the file        
NE.save ('TEST.xlsx')




