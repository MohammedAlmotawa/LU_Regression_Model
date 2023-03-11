#Defining (back to (CO2019) file for more details
import openpyxl
import os
from openpyxl.utils import get_column_letter
os.chdir("C:\\Users\\mohammedalmotawa\\Desktop\\PY_Master")

#Open a file in py and import CO data without the NULL DATA 

#The source Data
workbook = openpyxl.load_workbook('CO2019.xlsx')
sheet = workbook['Data']
#NE: new excel
NE = openpyxl.Workbook()
#Activate the sheet in the new excel file
NEws = NE.active
#Name the sheet
NEws.title = 'CO2019'
#Headings
NEws.append([ sheet ['A6'].value , sheet ['B4'].value , sheet ['C4'].value , sheet ['D4'].value , sheet ['E4'].value , sheet ['F4'].value , sheet ['G4'].value , sheet ['H4'].value , sheet ['I4'].value , sheet ['J4'].value ])



        
#Moving the rows that have all a real measurement of the CO concentration (without NULL values)
for row in range (7,50):
    for col in range (2, 11): #start from B
        char = get_column_letter(col)
        cell = char + str (row)
        celltype = type (sheet[cell].value)
        cellvalue = sheet[cell].value
        if celltype == float or celltype == int:
            NEws[cell] = cellvalue
        else:
            NEws.delete_rows(row)
            break

#Add the date
for row in range (7,50):
        cell = 'A' + str (row)
        NEws[cell]= str(sheet[cell].value)

   
#Delete empty rows
for row in range (7,50):
    for col in range (2,3):
        char = get_column_letter(col)
        cell = char + str (row)
        cellvalue = NEws[cell].value
        if cellvalue == '' or cellvalue == None:
            NEws.delete_rows(row)
            continue

##"%.2d" % day for day in range(1,24):


x = NEws['A7'].value
Day1 = NE.create_sheet ('Day1')
for row in range (7,50):
    char = get_column_letter (1)
    cell = char + str (row)
    x = NEws[cell].value
    if x.startswith("2019-01-01") or x.startswith ('01/01/2019'):
        Day1 [cell] = x
    continue
        
        
        

           

NE.save ('DataWithoutNULL.xlsx')




