#Defining (back to (Excel_CO2019) file for more details
import openpyxl
import os
from openpyxl.utils import get_column_letter
os.chdir("C:\\Users\\mmnnm\\Desktop\\PY_Master")

#Open a file in py and import CO data without the NULL DATA 

#The source Data
workbook = openpyxl.load_workbook('CO2019.xlsx')
sheet = workbook['Data']
#NE: new excel
NE = openpyxl.Workbook()
#Activate the sheet in the new excel file
NEws = NE.active
#Name the sheet
NEws.title = 'CO2019'
#Headings
NEws.append([ sheet ['A6'].value , sheet ['B4'].value , sheet ['C4'].value , sheet ['D4'].value , sheet ['E4'].value , sheet ['F4'].value , sheet ['G4'].value , sheet ['H4'].value , sheet ['I4'].value , sheet ['J4'].value ])



        
#Move the rows that have all a real measurment of the CO concentration (without NULL values)
for row in range (7,50):
    for col in range (2, 11): #start from B
        char = get_column_letter(col)
        cell = char + str (row)
        celltype = type (sheet[cell].value)
        cellvalue = sheet[cell].value
        if celltype == float or celltype == int:
            NEws[cell] = cellvalue
        else:
            NEws.delete_rows(row)
            break

#Add the date
for row in range (7,50):
        cell = 'A' + str (row)
        NEws[cell]= '\''+str(sheet[cell].value)

   
#Delete empty rows
for row in range (7,50):
    for col in range (2,3):
        char = get_column_letter(col)
        cell = char + str (row)
        cellvalue = NEws[cell].value
        if cellvalue == '' or cellvalue == None:
            NEws.delete_rows(row)
            continue



            
#Dividing the values based on the wieght
#FIRST STATION
##for i in range (7, 500):
##    for col in range (2,12):
##        char = chr (65 + str(col)
##        
##    Date = sheet.cell (row = i, column = 1).value
##    Date = sheet.cell (row = i, column = 1).value
##    if Station1 == 'NULL' or Station1 == '':
##        for x in range (2,8767):
##            if sheet1 ['A'+str(x)].value==None and sheet1 ['B'+str(x)].value==None:
##                sheet1['A'+str(x)] = Date
##                sheet1['B'+str(x)] = Station1
##                break



NE.save ('MoveTheDate_TheRealData_DeleteTheNULL.xlsx')
#iF THE DATE = THE OTHER DATES THEN THE DATA FOR THAT PERIOD ARE GOOD TO USE
#DELETE THE OTHER DATA THE HAVE NULL 




