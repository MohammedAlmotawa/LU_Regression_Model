#Defining (back to (Excel_CO2019) file for more details
import openpyxl
from openpyxl.utils import get_column_letter

#Open a file in py and import CO data without the NULL DATA 

#The source Data
workbook = openpyxl.load_workbook('CO2019.xlsx')
sheet = workbook['Data']
#NE: new excel
NE = openpyxl.Workbook()
#Activate the sheet in the new excel file
NEws = NE.active
#Name the sheet
NEws.title = 'CO2019'
#Headings
NEws.append([ sheet ['A1'].value , sheet ['B1'].value , sheet ['C1'].value , sheet ['D1'].value , sheet ['E1'].value , sheet ['F1'].value , sheet ['G1'].value , sheet ['H1'].value , sheet ['I1'].value , sheet ['J1'].value ])


#Add the date
for row in range (2,8759):
        cell = 'A' + str (row)
        NEws[cell]= '\''+str(sheet[cell].value)
        
#Move the rows that have all a real measurment of the CO concentration (without NULL values)
for row in range (2,8759):
    for col in range (2, 11): #start from B
        char = get_column_letter(col)
        cell = char + str (row)
        celltype = type (sheet[cell].value)
        cellvalue = sheet[cell].value
        if celltype == float or celltype == int:
            NEws[cell] = cellvalue
        else:
            break



   
#Delete empty rows
for row in range (2,8759):
    cell = 'J' + str (row)
    cellvalue = NEws[cell].value
    if cellvalue == '' or cellvalue is None:
        NEws.delete_rows(row)
    else:
        continue


NE.save ('CO_Step1_DATA_WithoutNULL.xlsx')






