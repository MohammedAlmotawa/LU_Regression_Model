#Defining (back to (Excel_CO2019) file for more details
import openpyxl
from openpyxl.utils import get_column_letter

#Open a file in py and import CO data without the NULL DATA 

#The source Data
workbook = openpyxl.load_workbook('CO_Step3_DeletedEmptyRows.xlsx')
sheet = workbook['CO2019']
#NE: new excel
NE = openpyxl.Workbook()

# Creating lists for days and months. Ex: List ['01','02',...]
ListDay = ["%.2d" % i for i in range(1,32)] 
ListMonth = ["%.2d" % i for i in range(1,13)]

#Creating sheets for each day
for D in range (1,32):
    NE.create_sheet(str(D))

#Add the same Headings as the original file for each sheeat
for d in range(1,32):
    SW1 = NE.active
    SW1 = NE[str(d)]
    SW1.append([ sheet ['A1'].value , sheet ['B1'].value , sheet ['C1'].value , sheet ['D1'].value , sheet ['E1'].value , sheet ['F1'].value , sheet ['G1'].value , sheet ['H1'].value , sheet ['I1'].value , sheet ['J1'].value ])


#Move every measurment based on the measurment month
for row in range (2,4172): #8766
    for m in range(12):
        for d in range (31):
            char = get_column_letter (1)
            cell = char + str (row)
            x = str(sheet[cell].value)
            month = ListMonth[m]
            day = ListDay[d]
            if x.startswith("'2019-"+month+"-"+day):
                SW2 = NE.active
                SW2 = NE[str(d+1)]
                SW2.append([sheet['A'+str(row)].value, sheet['B'+str(row)].value, sheet['C'+str(row)].value, sheet['D'+str(row)].value, sheet['E'+str(row)].value, sheet['F'+str(row)].value, sheet['G'+str(row)].value, sheet['H'+str(row)].value, sheet['I'+str(row)].value, sheet['J'+str(row)].value])
            else:
                continue
        

#Delete the first sheet
del NE['Sheet']

#Save the file        
NE.save ('CO_DividedByDay.xlsx')




