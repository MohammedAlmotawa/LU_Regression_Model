#Defining (back to (Excel_CO2019) file for more details
import openpyxl
import os
os.chdir("C:\\Users\\mmnnm\\Desktop\\PY_Master")





#1

#Show values from row 5 to row 10

##workbook = openpyxl.load_workbook('CO2019.xlsx')
##sheet = workbook['Data']
##for i in range (5, 10):
##    print (i, sheet.cell(row = i, column = 2).value)



#2
#Check every cell in column 2 wither it is a NULL or value < or > from 1
#OR above 2, and print the row number, the value, and a note

##workbook = openpyxl.load_workbook('CO2019.xlsx')
##sheet = workbook['Data']

##for i in range (7, 300):
##    value = sheet.cell (row = i, column = 2).value
##    if value == 'NULL':
##        print ('Null')
##    elif value < 1:
##        print (i, value,'Below One')
##    elif value >1 and value < 2:
##        print (i, value, 'Between 1 and 2')
##    else:
##        print (i, value,'Above 2')


#3
#Open a file in py and create different files for each group of values
#NULL, 1>,<1,>2)
#The source Data
workbook = openpyxl.load_workbook('CO2019.xlsx')
sheet = workbook['Data']
#NE: new excel
NE = openpyxl.Workbook()
sheet1 = NE.create_sheet('NULL')
sheet2 = NE.create_sheet('Less Than one')
sheet3 = NE.create_sheet('More Than one')
sheet4 = NE.create_sheet('More than two')
for i in range (7, 500):
    value = sheet.cell (row = i, column = 2).value
    Date = sheet.cell (row = i, column = 1).value
    if value == 'NULL' or value == '':
        for x in range (1,8766):
            if sheet1 ['B'+str(x)].value==None:
                sheet1['A'+str(x)] = Date
                sheet1['B'+str(x)] = value
                break

##            x = 1
##        if sheet1 ['B'+str(x)].value==None: 
##            sheet1['B'+str(x)] = value
##        elif sheet1 ['B'+str(x+1)].value ==None:
##            sheet1['B'+str(x+1)] = value
##        elif sheet1 ['B'+str(x+2)].value ==None:
##            sheet1['B'+str(x+2)] = value
##        elif sheet1 ['B'+str(x+3)].value ==None:
##            sheet1['B'+str(x+3)] = value
##        elif sheet1 ['B'+str(x+4)].value ==None:
##            sheet1['B'+str(x+4)] = value
##        elif sheet1 ['B'+str(x+5)].value ==None:
##            sheet1['B'+str(x+5)] = value
##        elif sheet1 ['B'+str(x+6)].value ==None:
##            sheet1['B'+str(x+6)] = value
##        elif sheet1 ['B'+str(x+7)].value ==None:
##            sheet1['B'+str(x+7)] = value
##        elif sheet1 ['B'+str(x+8)].value ==None:
##            sheet1['B'+str(x+8)] = value
##        elif sheet1 ['B'+str(x+9)].value ==None:
##            sheet1['B'+str(x+9)] = value
##        elif sheet1 ['B'+str(x+10)].value ==None:
##            sheet1['B'+str(x+10)] = value
##        elif sheet1 ['B'+str(x+11)].value ==None:
##            sheet1['B'+str(x+11)] = value
##        else:
##            sheet1['B'+str(x+12)] = value
    elif value < 1:
        for x in range (1,8766):
            if sheet2 ['B'+str(x)].value==None:
                sheet2['A'+str(x)] = Date
                sheet2['B'+str(x)] = value
                break
##        sheet2['B'+str(i)] = value
    elif value >1 and value < 2:
        for x in range (1,8766):
            if sheet3 ['B'+str(x)].value==None:
                sheet3['A'+str(x)] = Date
                sheet3['B'+str(x)] = value
                break
##        sheet3['B'+str(i)] = value
    else:
        for x in range (1,8766):
            if sheet4 ['B'+str(x)].value==None:
                sheet1['A'+str(x)] = Date
                sheet4['B'+str(x)] = value
                break
##        sheet4['B'+str(i)] = value
NE.save('test1.xlsx')






#Then maybe try to figure out the best way to find the peak records
#Also the the record that has a NULL as a value

