#Defining (back to (Excel_CO2019) file for more details
import openpyxl


#Open a file in py and create different files for each group of values
#NULL, 1>,<1,>2)
#The source Data
workbook = openpyxl.load_workbook('CO2019.xlsx')
sheet = workbook['Data']
#NE: new excel
NE = openpyxl.Workbook()
#Create sheets
sheet1 = NE.create_sheet('NULL')
sheet2 = NE.create_sheet('Less Than one')
sheet3 = NE.create_sheet('More Than one')
sheet4 = NE.create_sheet('More than two')

#Write the firs row with the same name as the original file
#First sheet
sheet1 = NE.active
sheet1 = NE['NULL']
sheet1.append([ sheet ['A1'].value , sheet ['B1'].value ,sheet ['A1'].value , sheet ['C1'].value ,sheet ['A1'].value , sheet ['D1'].value ,sheet ['A1'].value , sheet ['E1'].value ,sheet ['A1'].value , sheet ['F1'].value ,sheet ['A1'].value , sheet ['G1'].value ,sheet ['A1'].value , sheet ['H1'].value ,sheet ['A1'].value , sheet ['I1'].value ,sheet ['A1'].value , sheet ['J1'].value ])

#Second sheet
sheet2 = NE.active
sheet2 = NE['Less Than one']
sheet2.append([ sheet ['A1'].value , sheet ['B1'].value ,sheet ['A1'].value , sheet ['C1'].value ,sheet ['A1'].value , sheet ['D1'].value ,sheet ['A1'].value , sheet ['E1'].value ,sheet ['A1'].value , sheet ['F1'].value ,sheet ['A1'].value , sheet ['G1'].value ,sheet ['A1'].value , sheet ['H1'].value ,sheet ['A1'].value , sheet ['I1'].value ,sheet ['A1'].value , sheet ['J1'].value ])

#Third sheet
sheet3 = NE.active
sheet3 = NE['More Than one']
sheet3.append([ sheet ['A1'].value , sheet ['B1'].value ,sheet ['A1'].value , sheet ['C1'].value ,sheet ['A1'].value , sheet ['D1'].value ,sheet ['A1'].value , sheet ['E1'].value ,sheet ['A1'].value , sheet ['F1'].value ,sheet ['A1'].value , sheet ['G1'].value ,sheet ['A1'].value , sheet ['H1'].value ,sheet ['A1'].value , sheet ['I1'].value ,sheet ['A1'].value , sheet ['J1'].value ])

#Fourth sheet
sheet4 = NE.active
sheet4 = NE['More than two']
sheet4.append([ sheet ['A1'].value , sheet ['B1'].value ,sheet ['A1'].value , sheet ['C1'].value ,sheet ['A1'].value , sheet ['D1'].value ,sheet ['A1'].value , sheet ['E1'].value ,sheet ['A1'].value , sheet ['F1'].value ,sheet ['A1'].value , sheet ['G1'].value ,sheet ['A1'].value , sheet ['H1'].value ,sheet ['A1'].value , sheet ['I1'].value ,sheet ['A1'].value , sheet ['J1'].value ])

#Dividing the values based on the wieght
#FIRST STATION
for i in range (2, 8759): #8759
    Station1 = sheet.cell (row = i, column = 2).value
    Date = sheet.cell (row = i, column = 1).value
    if Station1 == 'NULL' or Station1 == '':
        for x in range (2,8759):
            if sheet1 ['A'+str(x)].value==None and sheet1 ['B'+str(x)].value==None:
                sheet1['A'+str(x)] = Date
                sheet1['B'+str(x)] = Station1
                break

    elif Station1 < 1:
        for x in range (2,8759):
            if sheet2 ['A'+str(x)].value==None and sheet2 ['B'+str(x)].value==None:
                sheet2['A'+str(x)] = Date
                sheet2['B'+str(x)] = Station1
                break

    elif Station1 < 2:
        for x in range (2,8759):
            if sheet3 ['A'+str(x)].value==None and sheet3 ['B'+str(x)].value==None:
                sheet3['A'+str(x)] = Date
                sheet3['B'+str(x)] = Station1
                break
    else:
        for x in range (2,8759):
            if sheet4 ['A'+str(x)].value==None and sheet4 ['B'+str(x)].value==None:
                sheet4['A'+str(x)] = Date
                sheet4['B'+str(x)] = Station1
                break
#SECOND STATION
for i in range (2, 8759):
    Station2 = sheet.cell (row = i, column = 3).value
    Date = sheet.cell (row = i, column = 1).value
    if Station2 == 'NULL' or Station2 == '':
        for x in range (2,8759):
            if sheet1 ['C'+str(x)].value==None and sheet1 ['D'+str(x)].value==None:
                sheet1['C'+str(x)] = Date
                sheet1['D'+str(x)] = Station2
                break

    elif Station2 < 1:
        for x in range (2,8759):
            if sheet2 ['C'+str(x)].value==None and sheet2 ['D'+str(x)].value==None:
                sheet2['C'+str(x)] = Date
                sheet2['D'+str(x)] = Station2
                break

    elif Station2 < 2:
        for x in range (2,8759):
            if sheet3 ['C'+str(x)].value==None and sheet3 ['D'+str(x)].value==None:
                sheet3['C'+str(x)] = Date
                sheet3['D'+str(x)] = Station2
                break
    else:
        for x in range (2,8759):
            if sheet4 ['C'+str(x)].value==None and sheet4 ['D'+str(x)].value==None:
                sheet4['C'+str(x)] = Date
                sheet4['D'+str(x)] = Station2
                break
#THIRD STATION
for i in range (2, 8759):
    Station3 = sheet.cell (row = i, column = 4).value
    Date = sheet.cell (row = i, column = 1).value
    if Station3 == 'NULL' or Station3 == '':
        for x in range (2,8759):
            if sheet1 ['E'+str(x)].value==None and sheet1 ['F'+str(x)].value==None:
                sheet1['E'+str(x)] = Date
                sheet1['F'+str(x)] = Station3
                break

    elif Station3 < 1:
        for x in range (2,8759):
            if sheet2 ['E'+str(x)].value==None and sheet2 ['F'+str(x)].value==None:
                sheet2['E'+str(x)] = Date
                sheet2['F'+str(x)] = Station3
                break

    elif Station3 < 2:
        for x in range (2,8759):
            if sheet3 ['E'+str(x)].value==None and sheet3 ['F'+str(x)].value==None:
                sheet3['E'+str(x)] = Date
                sheet3['F'+str(x)] = Station3
                break
    else:
        for x in range (2,8759):
            if sheet4 ['E'+str(x)].value==None and sheet4 ['F'+str(x)].value==None:
                sheet4['E'+str(x)] = Date
                sheet4['F'+str(x)] = Station3
                break
#4TH STATION
for i in range (2, 8759):
    Station4 = sheet.cell (row = i, column = 5).value
    Date = sheet.cell (row = i, column = 1).value
    if Station4 == 'NULL' or Station4 == '':
        for x in range (2,8759):
            if sheet1 ['G'+str(x)].value==None and sheet1 ['H'+str(x)].value==None:
                sheet1['G'+str(x)] = Date
                sheet1['H'+str(x)] = Station4
                break

    elif Station4 < 1:
        for x in range (2,8759):
            if sheet2 ['G'+str(x)].value==None and sheet2 ['H'+str(x)].value==None:
                sheet2['G'+str(x)] = Date
                sheet2['H'+str(x)] = Station4
                break

    elif Station4 < 2:
        for x in range (2,8759):
            if sheet3 ['G'+str(x)].value==None and sheet3 ['H'+str(x)].value==None:
                sheet3['G'+str(x)] = Date
                sheet3['H'+str(x)] = Station4
                break
    else:
        for x in range (2,8759):
            if sheet4 ['G'+str(x)].value==None and sheet4 ['H'+str(x)].value==None:
                sheet4['G'+str(x)] = Date
                sheet4['H'+str(x)] = Station4
                break
#5TH STATION
for i in range (2, 8759):
    Station5 = sheet.cell (row = i, column = 6).value
    Date = sheet.cell (row = i, column = 1).value
    if Station5 == 'NULL' or Station5 == '':
        for x in range (2,8759):
            if sheet1 ['I'+str(x)].value==None and sheet1 ['J'+str(x)].value==None:
                sheet1['I'+str(x)] = Date
                sheet1['J'+str(x)] = Station5
                break

    elif Station5 < 1:
        for x in range (2,8759):
            if sheet2 ['I'+str(x)].value==None and sheet2 ['J'+str(x)].value==None:
                sheet2['I'+str(x)] = Date
                sheet2['J'+str(x)] = Station5
                break

    elif Station5 < 2:
        for x in range (2,8759):
            if sheet3 ['I'+str(x)].value==None and sheet3 ['J'+str(x)].value==None:
                sheet3['I'+str(x)] = Date
                sheet3['J'+str(x)] = Station5
                break
    else:
        for x in range (2,8759):
            if sheet4 ['I'+str(x)].value==None and sheet4 ['J'+str(x)].value==None:
                sheet4['I'+str(x)] = Date
                sheet4['J'+str(x)] = Station5
                break
#6TH STATION
for i in range (2, 8759):
    Station6 = sheet.cell (row = i, column = 7).value
    Date = sheet.cell (row = i, column = 1).value
    if Station6 == 'NULL' or Station6 == '':
        for x in range (2,8759):
            if sheet1 ['K'+str(x)].value==None and sheet1 ['L'+str(x)].value==None:
                sheet1['K'+str(x)] = Date
                sheet1['L'+str(x)] = Station6
                break

    elif Station6 < 1:
        for x in range (2,8759):
            if sheet2 ['K'+str(x)].value==None and sheet2 ['L'+str(x)].value==None:
                sheet2['K'+str(x)] = Date
                sheet2['L'+str(x)] = Station6
                break

    elif Station6 < 2:
        for x in range (2,8759):
            if sheet3 ['K'+str(x)].value==None and sheet3 ['L'+str(x)].value==None:
                sheet3['K'+str(x)] = Date
                sheet3['L'+str(x)] = Station6
                break
    else:
        for x in range (2,8759):
            if sheet4 ['K'+str(x)].value==None and sheet4 ['L'+str(x)].value==None:
                sheet4['K'+str(x)] = Date
                sheet4['L'+str(x)] = Station6
                break
#7TH STATION
for i in range (2, 8759):
    Station7 = sheet.cell (row = i, column = 8).value
    Date = sheet.cell (row = i, column = 1).value
    if Station7 == 'NULL' or Station7 == '':
        for x in range (2,8759):
            if sheet1 ['M'+str(x)].value==None and sheet1 ['N'+str(x)].value==None:
                sheet1['M'+str(x)] = Date
                sheet1['N'+str(x)] = Station7
                break

    elif Station7 < 1:
        for x in range (2,8759):
            if sheet2 ['M'+str(x)].value==None and sheet2 ['N'+str(x)].value==None:
                sheet2['M'+str(x)] = Date
                sheet2['N'+str(x)] = Station7
                break

    elif Station7 < 2:
        for x in range (2,8759):
            if sheet3 ['M'+str(x)].value==None and sheet3 ['N'+str(x)].value==None:
                sheet3['M'+str(x)] = Date
                sheet3['N'+str(x)] = Station7
                break
    else:
        for x in range (2,8759):
            if sheet4 ['M'+str(x)].value==None and sheet4 ['N'+str(x)].value==None:
                sheet4['M'+str(x)] = Date
                sheet4['N'+str(x)] = Station7
                break
#8TH STATION
for i in range (2, 8759):
    Station8 = sheet.cell (row = i, column = 9).value
    Date = sheet.cell (row = i, column = 1).value
    if Station8 == 'NULL' or Station8 == '':
        for x in range (2,8759):
            if sheet1 ['O'+str(x)].value==None and sheet1 ['P'+str(x)].value==None:
                sheet1['O'+str(x)] = Date
                sheet1['P'+str(x)] = Station8
                break

    elif Station8 < 1:
        for x in range (2,8759):
            if sheet2 ['O'+str(x)].value==None and sheet2 ['P'+str(x)].value==None:
                sheet2['O'+str(x)] = Date
                sheet2['P'+str(x)] = Station8
                break

    elif Station8 < 2:
        for x in range (2,8759):
            if sheet3 ['O'+str(x)].value==None and sheet3 ['P'+str(x)].value==None:
                sheet3['O'+str(x)] = Date
                sheet3['P'+str(x)] = Station8
                break
    else:
        for x in range (2,8759):
            if sheet4 ['O'+str(x)].value==None and sheet4 ['P'+str(x)].value==None:
                sheet4['O'+str(x)] = Date
                sheet4['P'+str(x)] = Station8
                break
#9TH STATION
for i in range (2, 8759):
    Station9 = sheet.cell (row = i, column = 10).value
    Date = sheet.cell (row = i, column = 1).value
    if Station9 == 'NULL' or Station9 == '':
        for x in range (2,8759):
            if sheet1 ['Q'+str(x)].value==None and sheet1 ['R'+str(x)].value==None:
                sheet1['Q'+str(x)] = Date
                sheet1['R'+str(x)] = Station9
                break

    elif Station9 < 1:
        for x in range (2,8759):
            if sheet2 ['Q'+str(x)].value==None and sheet2 ['R'+str(x)].value==None:
                sheet2['Q'+str(x)] = Date
                sheet2['R'+str(x)] = Station9
                break

    elif Station9 < 2:
        for x in range (2,8759):
            if sheet3 ['Q'+str(x)].value==None and sheet3 ['R'+str(x)].value==None:
                sheet3['Q'+str(x)] = Date
                sheet3['R'+str(x)] = Station9
                break
    else:
        for x in range (2,8759):
            if sheet4 ['Q'+str(x)].value==None and sheet4 ['R'+str(x)].value==None:
                sheet4['Q'+str(x)] = Date
                sheet4['R'+str(x)] = Station9
                break

#Delete the first sheet
del NE['Sheet']

#Save the file
NE.save('CO_DividAllTheStationsValuesIntoFourCatagories.xlsx')



#iF THE DATE = THE OTHER DATES THEN THE DATA FOR THAT PERIOD ARE GOOD TO USE
#DELETE THE OTHER DATA THE HAVE NULL 




